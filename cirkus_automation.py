import csv
import json
import os
from datetime import datetime
import zipfile
import xml.etree.ElementTree as ET
import shutil

# Installation requise : pip install openpyxl
try:
    from openpyxl import load_workbook, Workbook
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    print("⚠️ openpyxl non installé. Installez-le avec : pip install openpyxl")

class CirkusOrderAutomation:
    def __init__(self):
        # Mapping entre les noms de produits de l'export et du bon de commande pour 10ML
        self.product_mapping_10ml = {
            # CLASSICS
            'CLASSIC FR - 10ML': 'CLASSIC FR',
            'CLASSIC RY4 - 10ML': 'CLASSIC RY4',
            'CLASSIC BLEND - 10ML': 'CLASSIC BLEND',
            'CLASSIC US - 10ML': 'CLASSIC US',
            'CLASSIC ORIGINAL - 10ML': 'CLASSIC ORIGINAL',
            'CLASSIC MENTHE - 10ML': 'CLASSIC MENTHE',
            'CLASSIC BLOND - 10ML': 'CLASSIC BLOND',
            'CLASSIC MENTHOL - 10ML': 'CLASSIC MENTHOL',
            'CLASSIC CERISE - 10ML': 'CLASSIC CERISE',
            'CLASSIC GOLD - 10ML': 'CLASSIC GOLD',
            'CLASSIC WHITE - 10ML': 'CLASSIC WHITE',
            
            # FRUITÉS
            'MANGUE FRAMBOISE - 10ML': 'MANGUE FRAMBOISE',
            'FRUITS ROUGES - 10ML': 'FRUITS ROUGES',
            'PASTEQUE MIX - 10ML': 'PASTEQUE MIX',
            'FRAMBOISE BLEUE - 10ML': 'FRAMBOISE BLEUE',
            'FRAISE KIWI - 10ML': 'FRAISE KIWI',
            'FRAMBOISE LITCHI - 10ML': 'FRAMBOISE LITCHI',
            'BONBON FRAISE - 10ML': 'BONBON FRAISE',
            'TROPICAL - 10ML': 'TROPICAL',
            'FRUIT DU DRAGON - 10ML': 'FRUIT DU DRAGON',
            'BONBON CERISE - 10ML': 'BONBON CERISE',
            'MANGUE PASSION VANILLE - 10ML': 'MANGUE PASSION VANILLE',
            'PINA FRAISE - 10ML': 'PINA FRAISE',
            'BONBON BANANE - 10ML': 'BONBON BANANE',
            
            # FRAIS
            'MENTHE POLAIRE - 10ML': 'MENTHE POLAIRE',
            'CASSIS FRAIS - 10ML': 'CASSIS FRAIS',
            'ABSINTHE ROUGE - 10ML': 'ABSINTHE ROUGE',
            'LEMON ICE - 10ML': 'LEMON ICE',
            'MENTHE CHLOROPHYLLE - 10ML': 'MENTHE CHLOROPHYLLE',
            'FRAISE MENTHE - 10ML': 'FRAISE MENTHE',
            
            # GIVRÉS
            'HANS LEGEL - 10ML': 'HANS LÉGEL (XTRA GIVRÉE)',
            'AL K\'POMME - 10ML': 'AL K\'POMME',
            'MURE A POINT - 10ML': 'MÛRE A POINT',
            'INST\'AGRUMES - 10ML': 'INST\'AGRUMES',
            'GARDE LA PECHE - 10ML': 'GARDE LA PÊCHE',
            'MANGUE DE SOLEIL - 10ML': 'MANGUE DE SOLEIL',
            'PRENDS LE MELON - 10ML': 'PRENDS LE MELON',
            'CASSIS CLAY - 10ML': 'CASSIS CLAY',
            'SODA RYAN - 10ML': 'SODA RYAN',
            
            # GOURMANDS
            'CARAMEL - 10ML': 'CARAMEL',
            'CAFE EXPRESSO - 10ML': 'CAFE EXPRESSO',
            'NOUGAT - 10ML': 'NOUGAT',
            'SWEET - 10ML': 'SWEET',
            'GOURMET - 10ML': 'GOURMET',
            'BRAVE - 10ML': 'BRAVE',
            'RESERVE - 10ML': 'RESERVE',
            'LOFTY - 10ML': 'LOFTY',
            'CHEESECAKE CITRON YUZU - 10ML': 'CHEESECAKE CITRON YUZU',
            'CACAHUETE CRUNCHY - 10ML': 'CACAHUETE CRUNCHY',
            'NOISETTE GOURMANDE - 10ML': 'NOISETTE GOURMANDE',
            'SAVAGE - 10ML': 'CLASSIC SAVAGE'
        }
        
        # Mapping pour les produits 50ML - PAB (Prêt à Booster)
        self.product_mapping_50ml = {
            # CLASSICS
            'CLASSIC FR - 50ML': 'CLASSIC FR',
            'CLASSIC RY4 - 50ML': 'CLASSIC RY4',
            'CLASSIC BLEND - 50ML': 'CLASSIC BLEND',
            'CLASSIC US - 50ML': 'CLASSIC US',
            'CLASSIC ORIGINAL - 50ML': 'CLASSIC ORIGINAL',
            'CLASSIC MENTHE - 50ML': 'CLASSIC MENTHE',
            'CLASSIC BLOND - 50ML': 'CLASSIC BLOND',
            'CLASSIC MENTHOL - 50ML': 'CLASSIC MENTHOL',
            'CLASSIC CERISE - 50ML': 'CLASSIC CERISE',
            'CLASSIC GOLD - 50ML': 'CLASSIC GOLD',
            'CLASSIC WHITE - 50ML': 'CLASSIC WHITE',
            
            # FRUITÉS
            'MANGUE FRAMBOISE - 50ML': 'MANGUE FRAMBOISE',
            'FRUITS ROUGES - 50ML': 'FRUITS ROUGES',
            'PASTEQUE MIX - 50ML': 'PASTEQUE MIX',
            'FRAMBOISE BLEUE - 50ML': 'FRAMBOISE BLEUE',
            'FRAISE KIWI - 50ML': 'FRAISE KIWI',
            'FRAMBOISE LITCHI - 50ML': 'FRAMBOISE LITCHI',
            'BONBON FRAISE - 50ML': 'BONBON FRAISE',
            'TROPICAL - 50ML': 'TROPICAL',
            'FRUIT DU DRAGON - 50ML': 'FRUIT DU DRAGON',
            'BONBON CERISE - 50ML': 'BONBON CERISE',
            'MANGUE PASSION VANILLE - 50ML': 'MANGUE PASSION VANILLE',
            'PINA FRAISE - 50ML': 'PINA FRAISE',
            'BONBON BANANE - 50ML': 'BONBON BANANE',
            
            # FRAIS
            'MENTHE POLAIRE - 50ML': 'MENTHE POLAIRE',
            'CASSIS FRAIS - 50ML': 'CASSIS FRAIS',
            'ABSINTHE ROUGE - 50ML': 'ABSINTHE ROUGE',
            'LEMON ICE - 50ML': 'LEMON ICE',
            'MENTHE CHLOROPHYLLE - 50ML': 'MENTHE CHLOROPHYLLE',
            'FRAISE MENTHE - 50ML': 'FRAISE MENTHE',
            
            # GIVRÉS
            'HANS LEGEL - 50ML': 'HANS LÉGEL (XTRA GIVRÉE)',
            'AL K\'POMME - 50ML': 'AL K\'POMME',
            'MURE A POINT - 50ML': 'MÛRE A POINT',
            'INST\'AGRUMES - 50ML': 'INST\'AGRUMES',
            'GARDE LA PECHE - 50ML': 'GARDE LA PÊCHE',
            'MANGUE DE SOLEIL - 50ML': 'MANGUE DE SOLEIL',
            'PRENDS LE MELON - 50ML': 'PRENDS LE MELON',
            'CASSIS CLAY - 50ML': 'CASSIS CLAY',
            'SODA RYAN - 50ML': 'SODA RYAN',
            
            # GOURMANDS
            'CARAMEL - 50ML': 'CARAMEL',
            'CAFE EXPRESSO - 50ML': 'CAFE EXPRESSO',
            'NOUGAT - 50ML': 'NOUGAT',
            'SWEET - 50ML': 'SWEET',
            'GOURMET - 50ML': 'GOURMET',
            'BRAVE - 50ML': 'BRAVE',
            'RESERVE - 50ML': 'RESERVE',
            'LOFTY - 50ML': 'LOFTY',
            'CHEESECAKE CITRON YUZU - 50ML': 'CHEESECAKE CITRON YUZU',
            'CACAHUETE CRUNCHY - 50ML': 'CACAHUETE CRUNCHY',
            'NOISETTE GOURMANDE - 50ML': 'NOISETTE GOURMANDE',
            'SAVAGE - 50ML': 'CLASSIC SAVAGE'
        }
        
        # Garder la compatibilité avec l'ancien code
        self.product_mapping = self.product_mapping_10ml
        
        self.template_path = None

    def set_template_path(self, template_path):
        """Set the Excel template path."""
        self.template_path = template_path

    def load_sales_data(self, file_path):
        """Load sales data from a CSV or XLSX file."""
        ext = os.path.splitext(file_path)[1].lower()
        data = []
        try:
            if ext == ".csv":
                with open(file_path, newline='', encoding='utf-8') as csvfile:
                    reader = csv.reader(csvfile)
                    data = [row for row in reader]
            elif ext in [".xlsx", ".xls"] and OPENPYXL_AVAILABLE:
                wb = load_workbook(file_path, read_only=True, data_only=True)
                ws = wb.active
                data = [[cell.value for cell in row] for row in ws.iter_rows()]
                wb.close()
            else:
                print(f"❌ Format de fichier non supporté ou openpyxl non installé : {file_path}")
                return None
            return data
        except Exception as e:
            print(f"❌ Erreur lors du chargement du fichier {file_path} : {e}")
            return None

    def get_client_list(self, clients_data):
        """Return a sorted list of client names."""
        if not clients_data:
            return []
        return sorted(clients_data.keys())

    def parse_client_selection(self, selection, clients):
        """Parse user selection input and return the list of selected clients."""
        selection = selection.strip()
        if not selection:
            return []
        if selection.lower() in ['all', 'tous', 'tout']:
            return clients
        if selection.lower() in ['exit', 'quitter']:
            return []
        selected = set()
        parts = [s.strip() for s in selection.split(',')]
        for part in parts:
            if part.isdigit():
                idx = int(part) - 1
                if 0 <= idx < len(clients):
                    selected.add(clients[idx])
            else:
                # Try to match by name (case-insensitive)
                for client in clients:
                    if part.lower() in client.lower():
                        selected.add(client)
        return list(selected)

    def generate_all_order_forms(self, selected_data):
        """Generate order forms for all selected clients."""
        for client, orders in selected_data.items():
            self.create_order_form_with_openpyxl_enhanced(client, orders)

    def fuzzy_match(self, a, b):
        """Basic fuzzy match: ignore accents and case, check if words are similar."""
        import unicodedata
        def normalize(s):
            return ''.join(
                c for c in unicodedata.normalize('NFD', s)
                if unicodedata.category(c) != 'Mn'
            ).lower()
        return normalize(a) == normalize(b)

    def parse_sales_data_by_format(self, raw_data, format_type="10ML"):
        """Parse les données de ventes selon le format (10ML ou 50ML)"""
        try:
            if not raw_data or len(raw_data) < 2:
                print("❌ Fichier vide ou invalide")
                return None
            
            # Déterminer le format à rechercher
            search_pattern = f"- {format_type}"
            
            # Trouver la ligne d'en-tête
            header_row = None
            data_start_row = None
            
            for i, row in enumerate(raw_data):
                if len(row) > 0 and any(search_pattern in str(cell) for cell in row[1:]):
                    header_row = i
                    data_start_row = i + 1
                    break
            
            if header_row is None:
                print(f"❌ Impossible de trouver les en-têtes de produits {format_type}")
                return None
            
            headers = [str(cell).strip() for cell in raw_data[header_row]]
            
            # Extraire les données clients
            clients_data = {}
            for row in raw_data[data_start_row:]:
                if len(row) > 0 and str(row[0]).strip():
                    client_name = str(row[0]).strip()
                    if client_name:
                        client_orders = {}
                        for j, value in enumerate(row[1:], 1):
                            if j < len(headers) and str(value).strip() and str(value).strip() != '0':
                                try:
                                    qty = int(float(str(value)))
                                    if qty > 0:
                                        product_name = headers[j]
                                        client_orders[product_name] = qty
                                except:
                                    pass
                        
                        if client_orders:  # Seulement si le client a des commandes
                            clients_data[client_name] = client_orders
            
            print(f"✅ Données {format_type} chargées : {len(clients_data)} clients trouvés")
            return clients_data
            
        except Exception as e:
            print(f"❌ Erreur lors du parsing {format_type} : {e}")
            return None

    def merge_client_data(self, clients_10ml, clients_50ml):
        """Fusionne les données des clients 10ML et 50ML"""
        merged_data = {}
        
        # Récupérer tous les clients uniques
        all_clients = set()
        if clients_10ml:
            all_clients.update(clients_10ml.keys())
        if clients_50ml:
            all_clients.update(clients_50ml.keys())
        
        for client in all_clients:
            merged_data[client] = {}
            
            # Ajouter les produits 10ML
            if clients_10ml and client in clients_10ml:
                merged_data[client].update(clients_10ml[client])
            
            # Ajouter les produits 50ML
            if clients_50ml and client in clients_50ml:
                merged_data[client].update(clients_50ml[client])
        
        print(f"✅ Données fusionnées : {len(merged_data)} clients au total")
        return merged_data

    def create_order_form_with_openpyxl_enhanced(self, client_name, client_orders, output_dir="bons_commande"):
        """Version améliorée qui gère les 10ML et 50ML"""
        if not OPENPYXL_AVAILABLE:
            print("❌ openpyxl requis pour conserver le formatage Excel")
            return self.create_order_form_csv_enhanced(client_name, client_orders, output_dir)
        
        if not self.template_path or not os.path.exists(self.template_path):
            print("❌ Template Excel non défini ou introuvable")
            return self.create_order_form_csv_enhanced(client_name, client_orders, output_dir)
        
        if not os.path.exists(output_dir):
            os.makedirs(output_dir)
        
        # Créer le nom du fichier
        date_str = datetime.now().strftime("%Y%m%d")
        filename = f"BON_COMMANDE_{client_name.replace(' ', '_').replace(',', '').replace('/', '_')}_{date_str}.xlsx"
        filepath = os.path.join(output_dir, filename)
        
        try:
            # Copier le template
            shutil.copy2(self.template_path, filepath)
            
            # Ouvrir et modifier le fichier
            workbook = load_workbook(filepath)
            sheet = workbook.active
            
            # Variables pour le suivi
            total_bottles_10ml = 0
            total_bottles_50ml = 0
            client_updated = False
            
            # Parcourir toutes les cellules pour trouver les éléments à modifier
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value is None:
                        continue
                    
                    cell_value = str(cell.value).strip()
                    
                    # 1. Mise à jour du nom du client
                    if not client_updated and (
                        cell_value.upper() == 'CLIENT' or 
                        'CLIENT' in cell_value.upper() or
                        cell_value.upper() == 'NOM CLIENT' or
                        'NOM' in cell_value.upper() and 'CLIENT' in cell_value.upper()
                    ):
                        next_cell = sheet.cell(row=cell.row, column=cell.column + 1)
                        if next_cell.value is None or str(next_cell.value).strip() == '':
                            next_cell.value = client_name
                            client_updated = True
                            print(f"✅ Client mis à jour : {client_name}")
                    
                    # 2. Mise à jour des quantités de produits
                    # 2. Mise à jour des quantités de produits
                    for export_product, quantity in client_orders.items():
                        template_product = None
                        is_50ml = False

                        # Déterminer le mapping à utiliser
                        if export_product in self.product_mapping_10ml:
                            template_product = self.product_mapping_10ml[export_product]
                            is_50ml = False
                        elif export_product in self.product_mapping_50ml:
                            template_product = self.product_mapping_50ml[export_product]
                            is_50ml = True

                        # Debug pour voir ce qui se passe
                        if template_product:
                            print(f"🔍 Debug: {export_product} -> {template_product} (50ML: {is_50ml})")
                        # Chercher le produit dans la feuille                        
                        if template_product and (template_product.upper() in cell_value.upper() or 
                                               cell_value.upper() in template_product.upper() or
                                               self.fuzzy_match(cell_value, template_product)):
                            
                            if is_50ml:
                                # Pour les 50ML, chercher la colonne PAB 50ML (position 11)
                                pab_cell = sheet.cell(row=cell.row, column=cell.column + 11)
                                if pab_cell.value is None or str(pab_cell.value).strip() in ['', '0']:
                                    pab_cell.value = quantity
                                    total_bottles_50ml += quantity
                                    print(f"✅ {template_product} (50ML): {quantity}")
                            else:
                                # Pour les 10ML, utiliser la colonne TOTAL (position 8)
                                qty_cell = sheet.cell(row=cell.row, column=cell.column + 8)
                                if qty_cell.value is None or str(qty_cell.value).strip() in ['', '0']:
                                    qty_cell.value = quantity
                                    total_bottles_10ml += quantity
                                    print(f"✅ {template_product} (10ML): {quantity}")
                            break
                    
                    # 3. Mise à jour des totaux
                    if ('TOTAL' in cell_value.upper() and 
                        ('FLACON' in cell_value.upper() or 'BOUTEILLE' in cell_value.upper())):
                        total_cell = sheet.cell(row=cell.row, column=cell.column + 1)
                        total_all = total_bottles_10ml + total_bottles_50ml
                        total_cell.value = total_all
                        print(f"✅ Total mis à jour : {total_all} (10ML: {total_bottles_10ml}, 50ML: {total_bottles_50ml})")
            
            # Sauvegarder le fichier
            workbook.save(filepath)
            workbook.close()
            
            print(f"✅ Bon de commande Excel créé : {filepath}")
            print(f"📦 Total flacons 10ML : {total_bottles_10ml}")
            print(f"📦 Total flacons 50ML : {total_bottles_50ml}")
            
            return filepath
            
        except Exception as e:
            print(f"❌ Erreur lors de la création avec openpyxl : {e}")
            return self.create_order_form_csv_enhanced(client_name, client_orders, output_dir)

    def create_order_form_csv_enhanced(self, client_name, client_orders, output_dir="bons_commande"):
        """Version CSV améliorée qui gère les 10ML et 50ML"""
        if not os.path.exists(output_dir):
            os.makedirs(output_dir)
        
        # Créer le nom du fichier
        date_str = datetime.now().strftime("%Y%m%d")
        filename = f"BON_COMMANDE_{client_name.replace(' ', '_').replace(',', '').replace('/', '_')}_{date_str}.csv"
        filepath = os.path.join(output_dir, filename)
        
        # Catégories et produits
        categories = {
            'CLASSICS': ['CLASSIC FR', 'CLASSIC RY4', 'CLASSIC BLEND', 'CLASSIC US',
                        'CLASSIC ORIGINAL', 'CLASSIC MENTHE', 'CLASSIC BLOND',
                        'CLASSIC MENTHOL', 'CLASSIC CERISE', 'CLASSIC GOLD', 'CLASSIC WHITE'],
            
            'FRUITÉS': ['MANGUE FRAMBOISE', 'FRUITS ROUGES', 'PASTEQUE MIX', 'FRAMBOISE BLEUE',
                       'FRAISE KIWI', 'FRAMBOISE LITCHI', 'PASSION', 'FRUITY PAMP\'',
                       'BONBON FRAISE', 'TROPICAL', 'FRUIT DU DRAGON', 'BONBON CERISE',
                       'MANGUE PASSION VANILLE', 'PINA FRAISE', 'BONBON BANANE'],
            
            'FRAIS': ['MENTHE POLAIRE', 'CASSIS FRAIS', 'ABSINTHE ROUGE', 'LEMON ICE',
                     'MENTHE CHLOROPHYLLE', 'FRAISE MENTHE', 'ABSINTHE POMME'],
            
            'GIVRÉS': ['HANS LÉGEL (XTRA GIVRÉE)', 'AL K\'POMME', 'MÛRE A POINT',
                      'INST\'AGRUMES', 'GARDE LA PÊCHE', 'MANGUE DE SOLEIL',
                      'PRENDS LE MELON', 'CASSIS CLAY', 'SODA RYAN'],
            
            'GOURMANDS': ['CARAMEL', 'CAFE EXPRESSO', 'NOUGAT', 'SWEET', 'GOURMET',
                         'BRAVE', 'RESERVE', 'LOFTY', 'CHEESECAKE CITRON YUZU',
                         'PECHE GOURMANDE', 'CACAHUETE CRUNCHY', 'VANILLE CUSTARD',
                         'CAFFE LATTE', 'NOISETTE GOURMANDE', 'CLASSIC SAVAGE']
        }
        
        total_bottles_10ml = 0
        total_bottles_50ml = 0
        
        try:
            with open(filepath, 'w', newline='', encoding='utf-8') as csvfile:
                writer = csv.writer(csvfile)
                
                # En-tête
                writer.writerow(['BON DE COMMANDE CIRKUS'])
                writer.writerow(['CLIENT', client_name])
                writer.writerow(['DIVALTO', ''])
                writer.writerow(['VILLE DE LIVRAISON', ''])
                writer.writerow([''])
                
                # Headers
                headers = ['SAVEUR', '0mg', '3mg', '6mg', '9mg', 'New Taux', '12mg', 
                    '16mg', 'TOTAL', 'SDN 10 mg', 'SDN 20 mg', 'PAB 50ML', 'AROMES 10mL', 'AROMES 30ml']
                writer.writerow(headers)
                
                # Catégories et produits
                for category, products in categories.items():
                    writer.writerow([category] + [''] * (len(headers) - 1))
                    
                    for product in products:
                        row = [product] + [0] * (len(headers) - 1)
                        
                        # Chercher le produit dans les commandes
                        for export_product, quantity in client_orders.items():
                            # Vérifier 10ML
                            if export_product in self.product_mapping_10ml:
                                if self.product_mapping_10ml[export_product] == product:
                                    row[8] = quantity  # Colonne TOTAL (position 8)
                                    total_bottles_10ml += quantity
                                    break
                            # Vérifier 50ML
                            elif export_product in self.product_mapping_50ml:
                                if self.product_mapping_50ml[export_product] == product:
                                    row[11] = quantity  # Colonne PAB 50ML (position 11)
                                    total_bottles_50ml += quantity
                                    break
                        
                        writer.writerow(row)
                
                # Total
                total_all = total_bottles_10ml + total_bottles_50ml
                writer.writerow(['TOTAL FLACONS', total_all] + [0] * (len(headers) - 2))
                writer.writerow(['  - dont 10ML', total_bottles_10ml] + [0] * (len(headers) - 2))
                writer.writerow(['  - dont 50ML', total_bottles_50ml] + [0] * (len(headers) - 2))
            
            print(f"✅ Bon de commande CSV créé : {filepath}")
            print(f"📦 Total flacons 10ML : {total_bottles_10ml}")
            print(f"📦 Total flacons 50ML : {total_bottles_50ml}")
            return filepath
            
        except Exception as e:
            print(f"❌ Erreur lors de la création du CSV : {e}")
            return None

    # Garder l'ancienne méthode mais rediriger vers la nouvelle
    def create_order_form_with_openpyxl(self, client_name, client_orders, output_dir="bons_commande"):
        return self.create_order_form_with_openpyxl_enhanced(client_name, client_orders, output_dir)

    def create_order_form_csv(self, client_name, client_orders, output_dir="bons_commande"):
        return self.create_order_form_csv_enhanced(client_name, client_orders, output_dir)

    # Garder la méthode parse_sales_data pour compatibilité
    def parse_sales_data(self, raw_data):
        return self.parse_sales_data_by_format(raw_data, "10ML")

    # Toutes les autres méthodes restent identiques...
    # [Le reste du code reste inchangé]

def main():
    """Fonction principale améliorée"""
    automation = CirkusOrderAutomation()
    
    print("🎯 AUTOMATISATION BON DE COMMANDE CIRKUS (10ML + 50ML)")
    print("=" * 60)
    
    # Vérifier openpyxl
    if not OPENPYXL_AVAILABLE:
        print("⚠️ ATTENTION : openpyxl n'est pas installé !")
        print("   Pour conserver le formatage Excel, installez-le avec :")
        print("   pip install openpyxl")
        print("")
    
    # Demander le template Excel
    print("📋 ÉTAPE 1 : Template de bon de commande")
    template_path = input("📁 Chemin vers votre template Excel (.xlsx) : ").strip().strip('"')
    
    if template_path and os.path.exists(template_path):
        automation.set_template_path(template_path)
    else:
        print("⚠️ Template non trouvé, utilisation du format CSV par défaut")
    
    # Demander le fichier 10ML
    print("\n📊 ÉTAPE 2 : Fichier d'export des ventes 10ML")
    sales_file_10ml = input("📁 Chemin vers votre fichier d'export 10ML (.xlsx ou .csv) : ").strip().strip('"')
    
    if not os.path.exists(sales_file_10ml):
        print("❌ Fichier 10ML non trouvé !")
        return
    
    # Demander le fichier 50ML
    print("\n📊 ÉTAPE 3 : Fichier d'export des ventes 50ML")
    print("💡 Laissez vide si vous n'avez pas de fichier 50ML")
    sales_file_50ml = input("📁 Chemin vers votre fichier d'export 50ML (.xlsx ou .csv) [optionnel] : ").strip().strip('"')
    
    # Charger les données 10ML
    print("🔄 Chargement des données 10ML...")
    raw_data_10ml = automation.load_sales_data(sales_file_10ml)
    if raw_data_10ml is None:
        return
    
    clients_data_10ml = automation.parse_sales_data_by_format(raw_data_10ml, "10ML")
    if clients_data_10ml is None:
        return
    
    # Charger les données 50ML si le fichier existe
    clients_data_50ml = None
    if sales_file_50ml and os.path.exists(sales_file_50ml):
        print("🔄 Chargement des données 50ML...")
        raw_data_50ml = automation.load_sales_data(sales_file_50ml)
        if raw_data_50ml:
            clients_data_50ml = automation.parse_sales_data_by_format(raw_data_50ml, "50ML")
    
    # Fusionner les données
    clients_data = automation.merge_client_data(clients_data_10ml, clients_data_50ml)
    clients = automation.get_client_list(clients_data)
    
    print(f"\n👥 {len(clients)} clients disponibles :")
    for i, client in enumerate(clients, 1):
        nb_products_10ml = len([p for p in clients_data[client].keys() if '10ML' in p]) if clients_data_10ml and client in clients_data_10ml else 0
        nb_products_50ml = len([p for p in clients_data[client].keys() if '50ML' in p]) if clients_data_50ml and client in clients_data_50ml else 0
        print(f"{i:2d}. {client} (10ML: {nb_products_10ml}, 50ML: {nb_products_50ml})")
    
    # Instructions pour la sélection
    print("\n" + "=" * 60)
    print("🎯 SÉLECTION DES CLIENTS :")
    print("   • Un seul client : entrez le numéro (ex: 3)")
    print("   • Plusieurs clients : séparez par des virgules (ex: 1,2,3 ou 1,4,6)")
    print("   • Vous pouvez aussi utiliser les noms de clients")
    print("   • Pour tous les clients : tapez 'all' ou 'tous'")
    print("   • Pour quitter : tapez 'exit' ou 'quitter'")
    
    selection = input("\n🎯 Votre sélection : ").strip()
    
    # Traitement de la sélection
    if selection.lower() in ['all', 'tous', 'tout']:
        selected_clients = clients
    else:
        selected_clients = automation.parse_client_selection(selection, clients)
    
    if not selected_clients:
        print("❌ Aucun client sélectionné ou sélection invalide")
        return
    
    # Génération des bons de commande
    print(f"\n🎯 {len(selected_clients)} client(s) sélectionné(s) :")
    for client in selected_clients:
        nb_products = len(clients_data[client])
        total_qty = sum(clients_data[client].values())
        print(f"   • {client} ({nb_products} produits, {total_qty} unités)")
    
    # Générer les bons de commande pour les clients sélectionnés
    selected_data = {client: clients_data[client] for client in selected_clients}
    automation.generate_all_order_forms(selected_data)
    
    print("\n✅ Traitement terminé !")

if __name__ == "__main__":
    main()